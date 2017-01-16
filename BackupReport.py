import pyodbc
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from math import ceil
import time

ROW = 4
Monday = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
          'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK',
          'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX']
Tuesday = ['AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN',
           'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ', 'CA', 'CB', 'CC',
           'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT']
Wednesday = ['CU', 'CV', 'CW', 'CX', 'CY', 'CZ', 'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN',
            'DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ', 'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EH',
            'EI', 'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP']
Thursday = ['EQ', 'ER', 'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ', 'FA', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FI', 'FJ',
            'FK', 'FL', 'FM', 'FN', 'FO', 'FP', 'FQ', 'FR', 'FS', 'FT', 'FU', 'FV', 'FW', 'FX', 'FY', 'FZ', 'GA', 'GB', 'GC', 'GD',
            'GE', 'GF', 'GG', 'GH', 'GI', 'GJ', 'GK', 'GL']
Friday = ['GM', 'GN', 'GO', 'GP', 'GQ', 'GR', 'GS', 'GT', 'GU', 'GV', 'GW', 'GX', 'GY', 'GZ', 'HA', 'HB', 'HC', 'HD', 'HE', 'HF', 'HG',
            'HH', 'HI', 'HJ', 'HK', 'HL', 'HM', 'HN', 'HO', 'HP', 'HQ', 'HR', 'HS', 'HT', 'HU', 'HV', 'HW', 'HX', 'HY', 'HZ', 'IA', 'IB',
            'IC', 'ID', 'IE', 'IF', 'IG', 'IH']
Saturday = ['II', 'IJ', 'IK', 'IL', 'IM', 'IN', 'IO', 'IP', 'IQ', 'IR', 'IS', 'IT', 'IU', 'IV', 'IW', 'IX', 'IY', 'IZ', 'JA', 'JB', 'JC',
            'JD', 'JE', 'JF', 'JG', 'JH','JI', 'JJ', 'JK', 'JL', 'JM', 'JN', 'JO', 'JP', 'JQ', 'JR', 'JS', 'JT', 'JU', 'JV', 'JW', 'JX',
            'JY', 'JZ', 'KA', 'KB', 'KC', 'KD']
Sunday = ['KE', 'KF', 'KG', 'KH', 'KI', 'KJ', 'KK', 'KL', 'KM', 'KN', 'KO', 'KP', 'KQ', 'KR', 'KS', 'KT', 'KU', 'KV', 'KW', 'KX', 'KY', 'KZ', 'LA', 'LB', 'LC', 'LD',
            'LE', 'LF', 'LG', 'LH', 'LI', 'LJ', 'LK', 'LL', 'LM', 'LN', 'LO', 'LP', 'LQ', 'LR', 'LS', 'LT', 'LU', 'LV', 'LW', 'LX', 'LY', 'LZ']
Server_Job_List = []
SQL_Request = '''
SELECT
     [sJOB].[name] AS [JobName],
     [freq_interval],
	 [active_start_time],
	 STUFF(RIGHT('000000' + CAST([active_start_time] AS VARCHAR(6)), 6), 3, 0,'')
	   as [Start_Time],
	STUFF(RIGHT('000000' + CAST([last_run_duration] AS VARCHAR(6)), 6), 3, 0,'')
	 as [last_run_duration]

FROM
    [msdb].[dbo].[sysjobs] AS [sJOB]
    LEFT JOIN [msdb].[sys].[servers] AS [sSVR]
        ON [sJOB].[originating_server_id] = [sSVR].[server_id]
    LEFT JOIN [msdb].[dbo].[sysjobschedules] AS [sJOBSCH]
        ON [sJOB].[job_id] = [sJOBSCH].[job_id]
    LEFT JOIN [msdb].[dbo].[sysschedules] AS [sSCH]
        ON [sJOBSCH].[schedule_id] = [sSCH].[schedule_id]
	LEFT JOIN [msdb].[dbo].[sysjobsteps] AS [sJSTEP]
		ON [sJOB].[job_id] = [sJSTEP].[job_id]
WHERE (sJOB.name LIKE 'Backup_FULL' OR sJOB.name LIKE 'Backup_DIFF') AND [sJSTEP].[step_id] = 1
'''


SQL_Servers = []
File_SQL_Servers = open('SQL_Servers.txt', 'r')
for line in File_SQL_Servers:
    SQL_Servers.append(line)



for Server in SQL_Servers:
    Server = Server.strip('\n')
    Connect = False
    try:
        ConnectStr = 'DRIVER={SQL Server};SERVER=' + Server + ';DATABASE=master'
        #	ConnectStr = 'DRIVER='+SQL_Data['Driver']+';SERVER='+SQL_Data['Server']+';DATABASE=master;UID='+SQL_Data['Sql_User']+';PWD='+SQL_Data['Password']
        SQL_connect = pyodbc.connect(ConnectStr + ';Trusted_Connection=yes')
        Connect = True
    except Exception, e:
        print Server
        print type(e), e
        print '_________'
    if not Connect:
        time.sleep(1)
        try:
            print Server
            ConnectStr = 'DRIVER={SQL Server};SERVER=' + Server + ';DATABASE=master;UID=zabbix;PWD=ZabbP@##'
            SQL_connect = pyodbc.connect(ConnectStr)
        except Exception, e:
            print type(e), e
    cursor = SQL_connect.cursor()
    cursor.execute(SQL_Request)
    row = cursor.fetchone()
    while row:
        Job = {}
        Job['Server'] = Server
        Job['JobName'] = row.JobName
        Job['freq_interval'] = row.freq_interval
        Job['Start_Time'] = row.Start_Time
        Job['last_run_duration'] = row.last_run_duration
        row = cursor.fetchone()
        Server_Job_List.append(Job)

GreenFill = PatternFill(start_color='FF00FF00',
                      end_color='FF00FF00',
                      fill_type='solid')
YellowFill = PatternFill(start_color='AAFF8000',
                      end_color='AAFF8000',
                      fill_type='solid')
xfile = openpyxl.load_workbook('report.xlsx')
sheet = xfile.get_sheet_by_name('report')

for Job in Server_Job_List:
    if Job['JobName'] == 'Backup_FULL':
        Fill = PatternFill(start_color='FF00FF00',
                      end_color='FF00FF00',
                      fill_type='solid')
    else:
        Fill = PatternFill(start_color='AAFF8000',
                      end_color='AAFF8000',
                      fill_type='solid')
    if Job['freq_interval'] & 2:
        sheet[('A'+str(ROW))] = Job['Server']
        sheet[('B'+str(ROW))] = Job['JobName']
        PointStart = ((int(Job['Start_Time'][:2])*3600)+(int(Job['Start_Time'][2:4])*60)+int(Job['Start_Time'][4:]))/1800
        LenPointStart = (((int(Job['last_run_duration'][:2])*3600)+(int(Job['last_run_duration'][2:4])*60)+(int(Job['last_run_duration'][4:])+1800))/1800)
        if ((int(Job['last_run_duration'])+1800)%1800) != 0:
            LenPointStart += 1
        i = 0
        ii = 0
        PointStart1 = 0
        print Job['Server'], Job['JobName'], Job['Start_Time'], PointStart, Job['last_run_duration'], LenPointStart
        while i != LenPointStart:
            if int(PointStart+i) > 47:
                sheet[(Tuesday[int(PointStart1+ii)]+str(ROW))] = ''
                sheet[(Tuesday[int(PointStart1+ii)]+str(ROW))].fill = Fill
                i+=1
                ii+=1
            else:
                sheet[(Monday[int(PointStart+i)]+str(ROW))] = ''
                sheet[(Monday[int(PointStart+i)]+str(ROW))].fill = Fill
                i+=1



    if Job['freq_interval'] & 4:
        sheet[('A'+str(ROW))] = Job['Server']
        sheet[('B'+str(ROW))] = Job['JobName']
        PointStart = ((int(Job['Start_Time'][:2])*3600)+(int(Job['Start_Time'][2:4])*60)+int(Job['Start_Time'][4:]))/1800
        LenPointStart = (((int(Job['last_run_duration'][:2])*3600)+(int(Job['last_run_duration'][2:4])*60)+(int(Job['last_run_duration'][4:])+1800))/1800)
        if ((int(Job['last_run_duration'])+1800)%1800) != 0:
            LenPointStart += 1
        i = 0
        ii = 0
        PointStart1 = 0
        while i != LenPointStart:
            if int(PointStart+i) > 47:
                sheet[(Wednesday[int(PointStart1+ii)]+str(ROW))] = ''
                sheet[(Wednesday[int(PointStart1+ii)]+str(ROW))].fill = Fill
                i+=1
                ii+=1
            else:
                sheet[(Tuesday[int(PointStart+i)]+str(ROW))] = ''
                sheet[(Tuesday[int(PointStart+i)]+str(ROW))].fill = Fill
                i+=1
    if Job['freq_interval'] & 8:
        sheet[('A'+str(ROW))] = Job['Server']
        sheet[('B'+str(ROW))] = Job['JobName']
        PointStart = ((int(Job['Start_Time'][:2])*3600)+(int(Job['Start_Time'][2:4])*60)+int(Job['Start_Time'][4:]))/1800
        LenPointStart = (((int(Job['last_run_duration'][:2])*3600)+(int(Job['last_run_duration'][2:4])*60)+(int(Job['last_run_duration'][4:])+1800))/1800)
        if ((int(Job['last_run_duration'])+1800)%1800) != 0:
            LenPointStart += 1
        i = 0
        ii = 0
        PointStart1 = 0
        while i != LenPointStart:
            if int(PointStart+i) > 47:
                sheet[(Thursday[int(PointStart1+ii)]+str(ROW))] = ''
                sheet[(Thursday[int(PointStart1+ii)]+str(ROW))].fill = Fill
                i+=1
                ii+=1
            else:
                sheet[(Wednesday[int(PointStart+i)]+str(ROW))] = ''
                sheet[(Wednesday[int(PointStart+i)]+str(ROW))].fill = Fill
                i+=1
    if Job['freq_interval'] & 16:
        sheet[('A'+str(ROW))] = Job['Server']
        sheet[('B'+str(ROW))] = Job['JobName']
        PointStart = ((int(Job['Start_Time'][:2])*3600)+(int(Job['Start_Time'][2:4])*60)+int(Job['Start_Time'][4:]))/1800
        LenPointStart = (((int(Job['last_run_duration'][:2])*3600)+(int(Job['last_run_duration'][2:4])*60)+(int(Job['last_run_duration'][4:])+1800))/1800)
        if ((int(Job['last_run_duration'])+1800)%1800) != 0:
            LenPointStart += 1
        i = 0
        ii = 0
        PointStart1 = 0
        while i != LenPointStart:
            if int(PointStart+i) > 47:
                sheet[(Friday[int(PointStart1+ii)]+str(ROW))] = ''
                sheet[(Friday[int(PointStart1+ii)]+str(ROW))].fill = Fill
                i+=1
                ii+=1
            else:
                sheet[(Thursday[int(PointStart+i)]+str(ROW))] = ''
                sheet[(Thursday[int(PointStart+i)]+str(ROW))].fill = Fill
                i+=1
    if Job['freq_interval'] & 32:
        sheet[('A'+str(ROW))] = Job['Server']
        sheet[('B'+str(ROW))] = Job['JobName']
        PointStart = ((int(Job['Start_Time'][:2])*3600)+(int(Job['Start_Time'][2:4])*60)+int(Job['Start_Time'][4:]))/1800
        LenPointStart = (((int(Job['last_run_duration'][:2])*3600)+(int(Job['last_run_duration'][2:4])*60)+(int(Job['last_run_duration'][4:])+1800))/1800)
        if ((int(Job['last_run_duration'])+1800)%1800) != 0:
            LenPointStart += 1
        i = 0
        ii = 0
        PointStart1 = 0
        while i != LenPointStart:
            if int(PointStart+i) > 47:
                sheet[(Saturday[int(PointStart1+ii)]+str(ROW))] = ''
                sheet[(Saturday[int(PointStart1+ii)]+str(ROW))].fill = Fill
                i+=1
                ii+=1
            else:
                sheet[(Friday[int(PointStart+i)]+str(ROW))] = ''
                sheet[(Friday[int(PointStart+i)]+str(ROW))].fill = Fill
                i+=1
    if Job['freq_interval'] & 64:
        sheet[('A'+str(ROW))] = Job['Server']
        sheet[('B'+str(ROW))] = Job['JobName']
        PointStart = ((int(Job['Start_Time'][:2])*3600)+(int(Job['Start_Time'][2:4])*60)+int(Job['Start_Time'][4:]))/1800
        LenPointStart = (((int(Job['last_run_duration'][:2])*3600)+(int(Job['last_run_duration'][2:4])*60)+(int(Job['last_run_duration'][4:])+1800))/1800)
        if ((int(Job['last_run_duration'])+1800)%1800) != 0:
            LenPointStart += 1
        i = 0
        ii = 0
        PointStart1 = 0
        while i != LenPointStart:
            if int(PointStart+i) > 47:
                sheet[(Sunday[int(PointStart1+ii)]+str(ROW))] = ''
                sheet[(Sunday[int(PointStart1+ii)]+str(ROW))].fill = Fill
                i+=1
                ii+=1
            else:
                sheet[(Saturday[int(PointStart+i)]+str(ROW))] = ''
                sheet[(Saturday[int(PointStart+i)]+str(ROW))].fill = Fill
                i+=1
    if Job['freq_interval'] & 1:
        sheet[('A'+str(ROW))] = Job['Server']
        sheet[('B'+str(ROW))] = Job['JobName']
        PointStart = ((int(Job['Start_Time'][:2])*3600)+(int(Job['Start_Time'][2:4])*60)+int(Job['Start_Time'][4:]))/1800
        LenPointStart = (((int(Job['last_run_duration'][:2])*3600)+(int(Job['last_run_duration'][2:4])*60)+(int(Job['last_run_duration'][4:])+1800))/1800)
        if ((int(Job['last_run_duration'])+1800)%1800) != 0:
            LenPointStart += 1
        i = 0
        ii = 0
        PointStart1 = 0
        while i != LenPointStart:
            if int(PointStart+i) > 47:
                sheet[(Monday[int(PointStart1+ii)]+str(ROW))] = ''
                sheet[(Monday[int(PointStart1+ii)]+str(ROW))].fill = Fill
                i+=1
                ii+=1
            else:
                sheet[(Sunday[int(PointStart+i)]+str(ROW))] = ''
                sheet[(Sunday[int(PointStart+i)]+str(ROW))].fill = Fill
                i+=1
    ROW += 1


xfile.save('report.xlsx')



